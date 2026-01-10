#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RECOLECTOR LÍNEA 141 - VERSIÓN GITHUB ACTIONS
Compatible con GitHub Actions + Local
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta
import pytz
import re
import time
import pandas as pd
import os
import argparse
from openpyxl.styles import Font
import json

TZ_AR = pytz.timezone('America/Argentina/Buenos_Aires')

PARADAS_INDIVIDUALES = [
    ("LP1912", "https://cuandollega.smartmovepro.net/nuevedejulio/arribos/?codLinea=141&idParada=LP1912"),
]

PARADAS_COMBINADAS = [
    ("L6203", "https://cuandollega.smartmovepro.net/nuevedejulio/arribos/?codLinea=141&idParada=L6203"),
    ("L6173", "https://cuandollega.smartmovepro.net/nuevedejulio/arribos/?codLinea=141&idParada=L6173"),
]

def get_fecha_excel():
    """Nombre del Excel de HOY: horarios-141-YYYY-MM-DD.xlsx"""
    return f"horarios-141-{datetime.now(TZ_AR).strftime('%Y-%m-%d')}.xlsx"

def minutos(texto: str):
    m = re.search(r'(\d+)\s*min', texto)
    return int(m.group(1)) if m else None

def scrape_parada(driver, nombre_parada, url):
    """Scraping de una parada"""
    try:
        driver.get(url)
        time.sleep(4)  # Esperar a que cargue el JSON

        # Ejecutar JavaScript para obtener el JSON de arribos
        try:
            arribos_json = driver.execute_script("return window._arribos;")
        except:
            arribos_json = []
        
        ahora = datetime.now(TZ_AR)
        horarios = []

        if arribos_json:
            # Si conseguimos el JSON, usarlo directamente
            for arribo in arribos_json:
                texto_tiempo = arribo.get('tiempoRestanteArribo', '')
                mins = minutos(texto_tiempo)
                if mins is None and "Arribando" in texto_tiempo:
                    mins = 0
                
                if mins is None:
                    continue
                
                hora = (ahora + timedelta(minutes=mins)).strftime("%H:%M")
                horarios.append({
                    'Hora_Scrap': ahora.strftime('%H:%M:%S'),
                    'Hora_Llegada': hora,
                    'Línea': arribo.get('descripcionBandera', ''),
                    'Minutos': mins,
                    'Parada': nombre_parada,
                    'CodigoColectivo': str(arribo.get('identificadorCoche', ''))
                })
        else:
            # Fallback: scraping del HTML si el JSON no está disponible
            tarjetas = driver.find_elements(By.CSS_SELECTOR, "div.mdl-grid.proximo-arribo")

            for card in tarjetas:
                try:
                    nombre_linea = card.find_element(
                        By.CSS_SELECTOR, "div.bandera h5"
                    ).text.strip()
                    texto_tiempo = card.find_element(
                        By.CSS_SELECTOR, "div.tiempo-arribo div"
                    ).text.strip()
                except Exception:
                    continue

                mins = minutos(texto_tiempo)
                if mins is None:
                    continue

                hora = (ahora + timedelta(minutes=mins)).strftime("%H:%M")
                horarios.append({
                    'Hora_Scrap': ahora.strftime('%H:%M:%S'),
                    'Hora_Llegada': hora,
                    'Línea': nombre_linea,
                    'Minutos': mins,
                    'Parada': nombre_parada,
                    'CodigoColectivo': ''
                })

        return horarios
    except Exception as e:
        print(f"   ❌ Error en {nombre_parada}: {e}")
        return []

def cargar_excel_dia():
    """Carga SOLO el Excel de HOY o crea vacío"""
    archivo_hoy = get_fecha_excel()
    
    if not os.path.exists(archivo_hoy):
        return {
            'LP1912': pd.DataFrame(),
            'LP1912-215': pd.DataFrame(),
            '6203-6173': pd.DataFrame()
        }
    
    try:
        excel_file = pd.ExcelFile(archivo_hoy)
        datos = {}
        
        for sheet in ['LP1912', 'LP1912-215', '6203-6173']:
            if sheet in excel_file.sheet_names:
                df = pd.read_excel(archivo_hoy, sheet_name=sheet, skiprows=4)
                datos[sheet] = df
            else:
                datos[sheet] = pd.DataFrame()
        
        return datos
    except Exception as e:
        print(f"⚠️ Error cargando {archivo_hoy}: {e}")
        return {
            'LP1912': pd.DataFrame(),
            'LP1912-215': pd.DataFrame(),
            '6203-6173': pd.DataFrame()
        }

def guardar_excel_dia(horarios_lp1912_nuevos, horarios_combinadas_nuevos):
    """Actualiza ONLY el Excel de HOY - SIN DUPLICADOS"""
    datos_existentes = cargar_excel_dia()
    ahora = datetime.now(TZ_AR)
    archivo_hoy = get_fecha_excel()
    
    # LP1912 principal
    df_lp1912_nuevos = pd.DataFrame(horarios_lp1912_nuevos)
    df_lp1912 = pd.concat([datos_existentes['LP1912'], df_lp1912_nuevos], ignore_index=True)
    df_lp1912 = df_lp1912.drop_duplicates(subset=['Hora_Llegada', 'Línea']).reset_index(drop=True)
    
    # Línea 215 separada
    nuevos_215 = [h for h in horarios_lp1912_nuevos if '215' in h.get('Línea', '')]
    df_215_nuevos = pd.DataFrame(nuevos_215)
    df_215 = pd.concat([datos_existentes['LP1912-215'], df_215_nuevos], ignore_index=True)
    df_215 = df_215.drop_duplicates(subset=['Hora_Llegada', 'Línea']).reset_index(drop=True)
    
    # Combinadas
    df_combinadas_nuevos = pd.DataFrame(horarios_combinadas_nuevos)
    df_combinadas = pd.concat([datos_existentes['6203-6173'], df_combinadas_nuevos], ignore_index=True)
    df_combinadas = df_combinadas.drop_duplicates(subset=['Hora_Llegada', 'Línea', 'Parada']).reset_index(drop=True)
    
    with pd.ExcelWriter(archivo_hoy, engine='openpyxl') as writer:
        df_lp1912.to_excel(writer, sheet_name='LP1912', index=False, startrow=4)
        df_215.to_excel(writer, sheet_name='LP1912-215', index=False, startrow=4)
        df_combinadas.to_excel(writer, sheet_name='6203-6173', index=False, startrow=4)
        
        sheets_info = {
            'LP1912': df_lp1912,
            'LP1912-215': df_215,
            '6203-6173': df_combinadas
        }
        
        for sheet_name, df in sheets_info.items():
            ws = writer.sheets[sheet_name]
            ws['A1'] = f'LÍNEA 141 - {sheet_name} - {ahora.strftime("%d/%m/%Y")}'
            ws['A2'] = f'Última actualización: {ahora.strftime("%H:%M:%S")}'
            ws['A3'] = f'Total filas: {len(df)}'
            
            for row in ws['A1:A3']:
                for cell in row:
                    cell.font = Font(bold=True)

    print(f"💾 Excel actualizado: {archivo_hoy}")
    print(f"   LP1912: {len(df_lp1912)} filas únicas")
    print(f"   215: {len(df_215)} filas únicas")
    print(f"   Combinadas: {len(df_combinadas)} filas únicas")

def guardar_txt(horarios, nombre_archivo, titulo):
    """Guarda horarios en archivo TXT"""
    ahora = datetime.now(TZ_AR)
    horarios_sorted = sorted(horarios, key=lambda x: x['Hora_Llegada'])
    
    with open(nombre_archivo, "w", encoding="utf-8") as f:
        f.write(f"🚌 {titulo}\n")
        f.write(f"📅 {ahora.strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"📊 {len(horarios_sorted)} horarios\n\n")
        for i, h in enumerate(horarios_sorted, 1):
            parada = h.get('Parada', '')
            codigo = h.get('CodigoColectivo', '')
            f.write(f"{i:2d}. {h['Hora_Llegada']} - {h['Línea']} ({h['Minutos']}min)")
            if codigo:
                f.write(f" [#{codigo}]")
            if parada:
                f.write(f" @ {parada}")
            f.write("\n")

def ciclo_completo(driver):
    """Un ciclo completo de scraping - GitHub Actions compatible"""
    ahora = datetime.now(TZ_AR)
    print(f"⏰ {ahora.strftime('%H:%M:%S')} - Recolectando datos...")
    t_inicio = time.time()
    
    # Scrapear LP1912
    print("   🌐 LP1912...", end=" ")
    horarios_lp1912 = scrape_parada(driver, "LP1912", PARADAS_INDIVIDUALES[0][1])
    print(f"✅ {len(horarios_lp1912)} buses")
    guardar_txt(horarios_lp1912, "horarios-LP1912.txt", "LÍNEA 141 - Parada LP1912")
    
    # Scrapear combinadas
    print("   🌐 Combinadas...", end=" ")
    horarios_combinadas = []
    for nombre, url in PARADAS_COMBINADAS:
        horarios = scrape_parada(driver, nombre, url)
        horarios_combinadas.extend(horarios)
    print(f"✅ {len(horarios_combinadas)} buses")
    guardar_txt(horarios_combinadas, "horarios-6203-6173.txt", "LÍNEA 141 - Paradas L6203 + L6173")
    
    # Actualizar Excel
    if horarios_lp1912 or horarios_combinadas:
        guardar_excel_dia(horarios_lp1912, horarios_combinadas)
    else:
        print("   ⚠️ Sin datos en esta ejecución")
    
    t_final = time.time()
    duracion = t_final - t_inicio
    print(f"✅ Ciclo completado en {duracion:.1f}s")

def main():
    parser = argparse.ArgumentParser(description="Recolector Línea 141")
    parser.add_argument('--once', action='store_true', help='Ejecutar 1 ciclo solamente (GitHub Actions)')
    args = parser.parse_args()
    
    # Chrome options OPTIMIZADAS para GitHub Actions
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-background-timer-throttling")
    chrome_options.add_argument("--disable-backgrounding-occluded-windows")
    chrome_options.add_argument("--disable-renderer-backgrounding")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    # Timeouts para GitHub
    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(25)
    driver.implicitly_wait(5)
    
    try:
        print("🚀 LÍNEA 141 - RECOLECTOR INICIADO")
        print(f"📅 {datetime.now(TZ_AR).strftime('%H:%M:%S')}")
        print(f"💻 Modo: {'GitHub Actions (1 ciclo)' if args.once else 'Local infinito'}")
        
        if args.once:
            ciclo_completo(driver)
        else:
            while True:
                ciclo_completo(driver)
                print("\n⏳ Espera: 15 minutos...\n")
                time.sleep(900)
                
    except KeyboardInterrupt:
        print("\n🛑 Programa detenido por usuario")
    except Exception as e:
        print(f"❌ Error general: {e}")
        raise
    finally:
        driver.quit()
        print("✅ Driver cerrado correctamente")

if __name__ == "__main__":
    main()
